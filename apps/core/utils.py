import os
from datetime import date, datetime
import openpyxl
from django.db import IntegrityError
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.db import transaction
from .models import (
    Product,
    AdditionalMasterData,
    BlockedProduct,
)

FILE_ADDITIONAL_PRODUCTS = "additional_products"
FILE_BLOCKED_PRODUCTS = "blocked_products"

REQUIRED_HEADERS = {
    FILE_ADDITIONAL_PRODUCTS: [
        "name",
        "articlenumber",
        "manufacturerpartnumber",
        "articlecalculationprices",
    ],
    FILE_BLOCKED_PRODUCTS: ["artikelnummer"],
}


def clean_text(value):
    return None if value in (None, "") else str(value).strip()


def normalize_header(s: str) -> str:
    """lowercase + remove non-alnum for robust header matching."""
    return "".join(ch for ch in str(s).strip().lower() if ch.isalnum())


def to_int(v, default: int = 1) -> int:
    try:
        i = int(float(v))
        return i if i >= 0 else default
    except Exception:
        return default


def to_bool(v, default=False):
    try:
        if str(v).strip().lower() in ("1", "true", "yes", "y", "j"):
            return True
        if str(v).strip().lower() in ("0", "false", "no", "n"):
            return False
        return default
    except Exception:
        return default


def to_decimal(v):
    if v is None or v == "":
        return None
    v = str(v).replace("€", "").strip()
    return round(float(v), 4)


def to_date(v) -> date | None:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def row_to_item(fields, file_type):
    """
    Convert normalized row fields to dict for DB insert/update.
    """

    if file_type == FILE_BLOCKED_PRODUCTS:
        name = clean_text(fields.get("name"))
        article_no = clean_text(fields.get("artikelnummer"))
        manufacturer_article_no = clean_text(fields.get("mpnnummer"))
        manufacturer = clean_text(fields.get("hersteller"))

        return {
            "name": name,
            "article_no": article_no,
            "manufacturer_article_no": manufacturer_article_no,
            "manufacturer": manufacturer,
        }

    name = clean_text(fields.get("name"))
    article_no = clean_text(fields.get("articlenumber"))
    description = clean_text(fields.get("itemdescription"))
    active = to_bool(fields.get("active"))
    width = to_decimal(fields.get("articlewidth"))
    height = to_decimal(fields.get("articleheight"))
    length = to_decimal(fields.get("articlelength"))
    weight = to_decimal(fields.get("articlenetweight"))
    manufacturer_article_no = clean_text(fields.get("manufacturerpartnumber"))
    article_calculation_price = to_decimal(fields.get("articlecalculationprices"))
    batch_number_required = to_bool(fields.get("batchnumberrequired"))
    gtin = clean_text(fields.get("gtin"))
    manufacturer = clean_text(fields.get("manufacturername"))
    stock = to_decimal(fields.get("stock"))
    store_refrigerated = to_bool(fields.get("storerefrigerated"), default=False)

    return {
        "name": name,
        "article_no": article_no,
        "description": description,
        "active": active,
        "width": width,
        "height": height,
        "length": length,
        "weight": weight,
        "manufacturer_article_no": manufacturer_article_no,
        "article_calculation_price": article_calculation_price,
        "batch_number_required": batch_number_required,
        "gtin": gtin,
        "manufacturer": manufacturer,
        "stock": stock,
        "store_refrigerated": store_refrigerated,
    }


def extract_rows_from_file(uploaded_file, file_type):
    """Reads first sheet of .xlsx and extracts the data as list of dict."""

    uploaded_file.seek(0)
    workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    # Get header row as plain values
    header_values = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)

    if not header_values:
        workbook.close()
        return []

    # Normalize headers
    headers = []
    for header_value in header_values:
        headers.append(normalize_header(header_value or ""))

    missing_columns = set(REQUIRED_HEADERS[file_type]) - set(headers)

    if missing_columns:
        workbook.close()
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    # Map each data row using the headers
    rows = []
    for values_row in sheet.iter_rows(min_row=3, values_only=True):
        if not values_row:
            continue

        row_has_data = False
        for cell in values_row:
            if isinstance(cell, str) and cell.strip():
                row_has_data = True
                break
            elif cell not in (None, ""):
                row_has_data = True
                break

        if not row_has_data:
            continue  # Skip empty row

        mapped_row = {}
        for index, header in enumerate(headers):
            mapped_row[header] = (
                values_row[index] if values_row and index < len(values_row) else None
            )
        rows.append(row_to_item(mapped_row, file_type))

    workbook.close()
    return rows


def validate_file_and_extract_rows(uploaded_file, file_type):
    name = getattr(uploaded_file, "name", "") or ""
    ext = os.path.splitext(name.lower())[1]
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        rows = extract_rows_from_file(uploaded_file, file_type)
        if not rows:
            raise ValueError("No data extracted")
        return rows
    raise ValueError("Unsupported file type (use .xlsx)")


# -------- DB upload --------


def upload_additional_products_to_db(rows, batch_size=500):
    unique_fields = ("manufacturer", "manufacturer_article_no", "name")

    # Extract keys as tuples
    keys = [
        tuple(row.get(f) for f in unique_fields)
        for row in rows
        if all(row.get(f) for f in unique_fields)
    ]

    # Build OR query for multiple keys
    q = Q()
    for key_tuple in keys:
        condition = Q()
        for field_name, value in zip(unique_fields, key_tuple):
            condition &= Q(**{field_name: value})
        q |= condition

    existing_qs = AdditionalMasterData.objects.filter(q)

    # Map existing objects using tuple key
    existing_map = {tuple(getattr(o, f) for f in unique_fields): o for o in existing_qs}

    # Build instances
    instances = []
    for row in rows:
        instances.append(
            AdditionalMasterData(
                name=row.get("name"),
                article_no=row.get("article_no"),
                description=row.get("description"),
                active=row.get("active"),
                width=row.get("width"),
                height=row.get("height"),
                length=row.get("length"),
                weight=row.get("weight"),
                manufacturer_article_no=row.get("manufacturer_article_no"),
                article_calculation_price=row.get("article_calculation_price"),
                batch_number_required=row.get("batch_number_required"),
                gtin=row.get("gtin"),
                manufacturer=row.get("manufacturer"),
                stock=row.get("stock"),
                store_refrigerated=row.get("store_refrigerated"),
            )
        )

    if not instances:
        return {"created": 0, "updated": 0}

    fields = [
        "name",
        "description",
        "active",
        "width",
        "height",
        "length",
        "weight",
        "manufacturer_article_no",
        "article_calculation_price",
        "batch_number_required",
        "gtin",
        "manufacturer",
        "stock",
        "store_refrigerated",
    ]

    created_count, updated_count = 0, 0

    for i in range(0, len(instances), batch_size):
        batch = instances[i : i + batch_size]
        to_update = []
        to_create = []

        for obj in batch:
            key = tuple(getattr(obj, f) for f in unique_fields)
            if key in existing_map:
                existing_obj = existing_map[key]
                for f in fields:
                    setattr(existing_obj, f, getattr(obj, f))
                to_update.append(existing_obj)
            else:
                to_create.append(obj)

        with transaction.atomic():
            try:
                if to_update:
                    AdditionalMasterData.objects.bulk_update(to_update, fields)
                    updated_count += len(to_update)
                if to_create:
                    AdditionalMasterData.objects.bulk_create(to_create)
                    created_count += len(to_create)

            except IntegrityError:
                raise Exception(
                    "Duplicate Manufacturer + Manufacturer Article Number detected. "
                    "Please clean up your data and try again."
                )

    return {
        "created": created_count,
        "updated": updated_count,
        "total": created_count + updated_count,
    }


def upload_blocked_products_to_db(rows, batch_size=500):
    unique_fields = ("article_no", "manufacturer_article_no")

    # Build tuple keys
    keys = [
        tuple(row.get(f) for f in unique_fields)
        for row in rows
        if all(row.get(f) for f in unique_fields)
    ]

    # Build OR query
    q = Q()
    for a, b in keys:
        q |= Q(article_no=a, manufacturer_article_no=b)

    existing_qs = BlockedProduct.objects.filter(q)

    # Map existing by tuple key
    existing_map = {(o.article_no, o.manufacturer_article_no): o for o in existing_qs}

    # Build instances
    instances = [
        BlockedProduct(
            name=row.get("name"),
            article_no=row.get("article_no"),
            manufacturer_article_no=row.get("manufacturer_article_no"),
            manufacturer=row.get("manufacturer"),
        )
        for row in rows
    ]

    if not instances:
        return {"created": 0, "updated": 0}

    fields = ["name", "manufacturer_article_no", "manufacturer"]

    created_count, updated_count = 0, 0

    # Batch processing
    for i in range(0, len(instances), batch_size):
        batch = instances[i : i + batch_size]
        to_update, to_create = [], []

        for obj in batch:
            key = (obj.article_no, obj.manufacturer_article_no)
            if key in existing_map:
                existing_obj = existing_map[key]
                for f in fields:
                    setattr(existing_obj, f, getattr(obj, f))
                to_update.append(existing_obj)
            else:
                to_create.append(obj)

        with transaction.atomic():

            try:
                if to_update:
                    BlockedProduct.objects.bulk_update(to_update, fields)
                    updated_count += len(to_update)

                if to_create:
                    BlockedProduct.objects.bulk_create(to_create)
                    created_count += len(to_create)
            except IntegrityError:
                raise Exception(
                    "Duplicate Manufacturer + Manufacturer Article Number detected. "
                    "Please clean up your data and try again."
                )

    return {
        "created": created_count,
        "updated": updated_count,
        "total": created_count + updated_count,
    }


def sync_product_relations(
    model, product_field="product", has_sku=True, is_gls_model=True
):
    product_map = dict(Product.objects.values_list("sku", "id"))

    objs = model.objects.filter(**{f"{product_field}__isnull": True})

    updates = []

    if is_gls_model:
        for obj in objs.only("id", "article_no"):
            sku = f"LG{obj.article_no}" if obj.article_no else None
            product_id = product_map.get(sku)
            if product_id:
                updates.append(model(id=obj.id, **{f"{product_field}_id": product_id}))

    elif not is_gls_model and not has_sku:
        for obj in objs.only("id", "article_no", "manufacturer"):
            sku = (
                f"{(obj.manufacturer or '')[:2].upper()}{obj.article_no}"
                if obj.article_no
                else None
            )
            product_id = product_map.get(sku)
            if product_id:
                updates.append(model(id=obj.id, **{f"{product_field}_id": product_id}))
    else:
        for obj in objs.only("id", "sku"):
            sku = obj.sku
            product_id = product_map.get(sku)
            if product_id:
                updates.append(model(id=obj.id, **{f"{product_field}_id": product_id}))

    if updates:
        model.objects.bulk_update(updates, [f"{product_field}_id"], batch_size=1000)

    return True
