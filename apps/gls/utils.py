from apps.core.models import LogEntry
import os
import re
from apps.gls.mapping import field_map_101
from apps.gls.mapping import field_map_102
from django.conf import settings
import os
from datetime import date, datetime
import openpyxl
from django.db.models import Q
from django.db import transaction
from .models import (
    GLSProductGroup,
)

GLS_UPLOAD_PATH = settings.GLS_UPLOAD_PATH
FILE_PRODUCT_GROUP = "product_group"

REQUIRED_HEADERS = {
    FILE_PRODUCT_GROUP: [
        "code",
        "beschreibung",
    ],
}


class GlsLog:
    @staticmethod
    def info(msg):
        LogEntry.objects.create(source=LogEntry.GLS, level=LogEntry.INFO, message=msg)

    @staticmethod
    def warning(msg):
        LogEntry.objects.create(
            source=LogEntry.GLS, level=LogEntry.WARNING, message=msg
        )

    @staticmethod
    def error(msg):
        LogEntry.objects.create(source=LogEntry.GLS, level=LogEntry.ERROR, message=msg)


def export_gls_orders_to_csv(order_header, delimiter="^#!"):
    header_rows = []
    order_lines = []

    header_fields = field_map_101["fields"]
    header_values = [
        _format_gls_value(getattr(order_header, field_name), field_type)
        for field_name, field_type in header_fields
    ]
    header_rows.append(delimiter.join(map(str, header_values)))

    line_fields = field_map_102["fields"]
    for line in order_header.lines.all():
        line_values = [
            _format_gls_value(getattr(line, field_name), field_type)
            for field_name, field_type in line_fields
        ]
        order_lines.append(delimiter.join(map(str, line_values)))

    os.makedirs(GLS_UPLOAD_PATH, exist_ok=True)
    header_csv_temp_name = f"{order_header.order_no}.101.tmp"
    header_csv_perm_name = f"{order_header.order_no}.101"

    lines_csv_temp_name = f"{order_header.order_no}.102.tmp"
    lines_csv_perm_name = f"{order_header.order_no}.102"

    header_csv_path = os.path.join(GLS_UPLOAD_PATH, header_csv_perm_name)
    lines_csv_path = os.path.join(GLS_UPLOAD_PATH, lines_csv_perm_name)

    header_data = "\r\n".join(header_rows)
    lines_data = "\r\n".join(order_lines)

    with open(header_csv_path, "w", encoding="cp850") as f:
        f.write(header_data)
    with open(lines_csv_path, "w", encoding="cp850") as f:
        f.write(lines_data)

    return {
        "header_csv_path": header_csv_path,
        "header_csv_temp_name": header_csv_temp_name,
        "header_csv_perm_name": header_csv_perm_name,
        "lines_csv_path": lines_csv_path,
        "lines_csv_temp_name": lines_csv_temp_name,
        "lines_csv_perm_name": lines_csv_perm_name,
    }


def _format_gls_value(value, field_type):
    if value in (None, ""):
        return ""

    match = re.match(r"(\d+)?([a-zA-Z,/_]+)", field_type)
    length = int(match.group(1)) if match.group(1) else None
    ftype = match.group(2)

    if ftype.endswith("c"):
        value = str(value).strip().upper()
    elif ftype.endswith("t"):
        value = str(value).strip()
    elif ftype == "d":
        value = str(int(float(value))).replace(".", "").replace(",", "")
    elif ftype == "d,":
        value = str(value).replace(",", "").replace(".", ",")
    elif ftype.startswith("TT"):
        value = value.strftime("%d.%m.%y") if value else ""
    elif ftype in ("J/N", "c_bool"):
        value = "J" if value else "N"
    else:
        value = str(value)

    if length and not ftype.startswith("TT"):
        value = value[:length]

    return value


def clean_text(value):
    return None if value in (None, "") else str(value).strip()


def normalize_header(s: str) -> str:
    """lowercase + remove non-alnum for robust header matching."""
    return "".join(ch for ch in str(s).strip().lower() if ch.isalnum())


def to_int(v, default: int = 1) -> int:
    try:
        i = int(float(v))
        return i if i >= 0 else default
    except Exception:
        return default


def to_bool(v, default=False):
    try:
        if str(v).strip().lower() in ("1", "true", "yes", "y", "j"):
            return True
        if str(v).strip().lower() in ("0", "false", "no", "n"):
            return False
        return default
    except Exception:
        return default


def to_decimal(v):
    if v is None or v == "":
        return None
    v = str(v).replace("€", "").strip()
    return round(float(v), 4)


def to_date(v) -> date | None:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def row_to_item(fields, file_type):
    """
    Convert normalized row fields to dict for DB insert/update.
    """

    if file_type == FILE_PRODUCT_GROUP:
        product_group_no = clean_text(fields.get("code"))
        product_group_name = clean_text(fields.get("beschreibung"))

        return {
            "product_group_no": product_group_no,
            "product_group_name": product_group_name,
        }


def extract_rows_from_file(uploaded_file, file_type):
    """Reads first sheet of .xlsx and extracts the data as list of dict."""

    uploaded_file.seek(0)
    workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    # Get header row as plain values
    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)

    if not header_values:
        workbook.close()
        return []

    # Normalize headers
    headers = []
    for header_value in header_values:
        headers.append(normalize_header(header_value or ""))

    missing_columns = set(REQUIRED_HEADERS[file_type]) - set(headers)

    if missing_columns:
        workbook.close()
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    # Map each data row using the headers
    rows = []
    for values_row in sheet.iter_rows(min_row=2, values_only=True):
        if not values_row:
            continue

        row_has_data = False
        for cell in values_row:
            if isinstance(cell, str) and cell.strip():
                row_has_data = True
                break
            elif cell not in (None, ""):
                row_has_data = True
                break

        if not row_has_data:
            continue  # Skip empty row

        mapped_row = {}
        for index, header in enumerate(headers):
            mapped_row[header] = (
                values_row[index] if values_row and index < len(values_row) else None
            )
        rows.append(row_to_item(mapped_row, file_type))

    workbook.close()
    return rows


def validate_file_and_extract_rows(uploaded_file, file_type):
    name = getattr(uploaded_file, "name", "") or ""
    ext = os.path.splitext(name.lower())[1]
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        rows = extract_rows_from_file(uploaded_file, file_type)
        if not rows:
            raise ValueError("No data extracted")
        return rows
    raise ValueError("Unsupported file type (use .xlsx)")


def upload_product_group_to_db(rows, batch_size=500):
    unique_fields = ("product_group_no",)

    # Extract keys based on unique_fields
    keys = [
        tuple(row.get(f) for f in unique_fields)
        for row in rows
        if all(row.get(f) for f in unique_fields)
    ]

    # Build OR query dynamically
    q = Q()
    for key in keys:
        q |= Q(**{field: value for field, value in zip(unique_fields, key)})

    existing_qs = GLSProductGroup.objects.filter(q)

    # Map existing objects by their unique tuple
    existing_map = {tuple(getattr(o, f) for f in unique_fields): o for o in existing_qs}

    # Build new objects
    instances = [
        GLSProductGroup(
            product_group_no=row.get("product_group_no"),
            product_group_name=row.get("product_group_name"),
        )
        for row in rows
    ]

    if not instances:
        return {"created": 0, "updated": 0}

    update_fields = ["product_group_name"]
    created_count = updated_count = 0

    for i in range(0, len(instances), batch_size):
        batch = instances[i : i + batch_size]
        to_create, to_update = [], []

        for obj in batch:
            key = tuple(getattr(obj, f) for f in unique_fields)
            if key in existing_map:
                existing_obj = existing_map[key]
                for f in update_fields:
                    setattr(existing_obj, f, getattr(obj, f))
                to_update.append(existing_obj)
            else:
                to_create.append(obj)

        with transaction.atomic():
            if to_update:
                GLSProductGroup.objects.bulk_update(to_update, update_fields)
                updated_count += len(to_update)

            if to_create:
                GLSProductGroup.objects.bulk_create(to_create)
                created_count += len(to_create)

    return {
        "created": created_count,
        "updated": updated_count,
        "total": created_count + updated_count,
    }
